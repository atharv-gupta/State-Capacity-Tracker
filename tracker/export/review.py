#!/usr/bin/env python3
"""Export the federal and congressional trackers to review workbooks.

Airtable is where review annotations live (`review_status` / `reviewer_notes`,
which the upsert path preserves on every nightly run), but Airtable is a poor
surface for a colleague who has been asked to read 80 rows and mark them up.
This writes one .xlsx per tracker, filterable and sortable, with the review
columns already in place and a Read me sheet that states what the reviewer is
being asked to decide.

    python export_review.py                            # both, last 21 days
    python export_review.py --tracker federal --days 7
    python export_review.py --all --out ~/Desktop
    python export_review.py --relevant-only            # drop the `none` rows

Whatever reviewers write in the workbook has to be typed back into Airtable's
review_status / reviewer_notes to survive — the export is one-way on purpose.
Round-tripping would mean reconciling two sources of truth for the same field.
"""

import argparse
import os
import sys
from datetime import date, timedelta

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pyairtable import Api

load_dotenv()

AIRTABLE_TOKEN = os.environ.get("AIRTABLE_TOKEN")
AIRTABLE_BASE_ID = os.environ.get("AIRTABLE_BASE_ID")

# Columns the reviewer fills in. Appended to every sheet, with a dropdown on
# `verdict` so the values stay countable.
REVIEW_COLUMNS = ["verdict", "correction", "reviewer notes", "reviewer"]
VERDICT_CHOICES = ["keep", "drop — not an action", "drop — not capacity",
                   "wrong competency", "wrong relevance", "needs edit"]

HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
REVIEW_FILL = PatternFill("solid", fgColor="FEF3C7")
NONE_FILL = PatternFill("solid", fgColor="F8FAFC")


def joined(value):
    """Airtable multi-selects arrive as lists; multiline text as one string."""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return (value or "").replace("\n", " · ") if isinstance(value, str) else value


def first_url(value):
    for line in (value or "").splitlines():
        if line.strip():
            return line.strip()
    return ""


# ---------------------------------------------------------------------------
# Sheet specs. Each column is (header, airtable field, width). The order is the
# reading order for a reviewer: when, who, what kind, our classification, then
# the prose, then the evidence.
# ---------------------------------------------------------------------------
FEDERAL_EVENTS = {
    "table": "Federal Events",
    "sheet": "Federal events",
    "date_field": "date",
    "columns": [
        ("date", "date", 11),
        ("lane", "lane", 16),
        ("branch", "branch", 10),
        ("agency", "agency", 18),
        ("instrument", "instrument_type", 20),
        ("instrument id", "instrument_id", 14),
        ("verification", "verification", 13),
        ("competency", "competency", 22),
        ("relevance", "relevance", 9),
        ("topic tags", "topic_tags", 24),
        ("title", "short_title", 46),
        ("headline", "headline", 60),
        ("summary", "summary", 70),
        ("why it matters", "why_it_matters", 50),
        ("actor", "actor", 22),
        ("status", "status", 16),
        ("sources", "article_count", 8),
        ("outlets", "source_outlets", 30),
        ("link", "source_urls", 40),
        ("primary document", "document_url", 40),
        ("airtable review_status", "review_status", 16),
    ],
}

CONGRESS_EVENTS = {
    "table": "Congress Events",
    "sheet": "Committee activity",
    "date_field": "date",
    "columns": [
        ("date", "date", 11),
        ("committee", "committee", 16),
        ("chamber", "chamber", 9),
        ("party source", "party_source", 12),
        ("activity", "activity_type", 18),
        ("competency", "competency", 22),
        ("relevance", "relevance", 9),
        ("topic tags", "topic_tags", 24),
        ("title", "short_title", 46),
        ("headline", "headline", 60),
        ("summary", "summary", 70),
        ("why it matters", "why_it_matters", 50),
        ("actor", "actor", 22),
        ("bills", "bill_refs", 16),
        ("status", "status", 14),
        ("sources", "article_count", 8),
        ("outlets", "source_outlets", 30),
        ("link", "source_urls", 40),
        ("airtable review_status", "review_status", 16),
    ],
}

CONGRESS_HEARINGS = {
    "table": "Congress Hearings",
    "sheet": "Hearings",
    "date_field": "date",
    "columns": [
        ("date", "date", 11),
        ("committee", "committee", 16),
        ("chamber", "chamber", 9),
        ("status", "hearing_status", 11),
        ("type", "meeting_type", 14),
        ("competency", "competency", 22),
        ("relevance", "relevance", 9),
        ("topic tags", "topic_tags", 24),
        ("title", "short_title", 50),
        ("agenda summary", "agenda_summary", 70),
        ("why it matters", "why_it_matters", 50),
        ("location", "location", 20),
        ("bills", "bill_refs", 20),
        ("link", "source_urls", 40),
        ("airtable review_status", "review_status", 16),
    ],
}

CONGRESS_BILLS = {
    "table": "Congress Bills",
    "sheet": "Bills",
    "date_field": "date",
    "columns": [
        ("committee action date", "committee_action_date", 13),
        ("bill", "bill_number", 12),
        ("committee", "committee", 16),
        ("chamber", "chamber", 9),
        ("committee action", "committee_action", 15),
        ("bill status", "bill_status", 14),
        ("competency", "competency", 22),
        ("relevance", "relevance", 9),
        ("topic tags", "topic_tags", 24),
        ("title", "title", 60),
        ("summary", "summary", 70),
        ("why it matters", "why_it_matters", 50),
        ("sponsor", "sponsor", 22),
        ("cosponsors", "cosponsor_count", 10),
        ("latest action", "latest_action", 40),
        ("link", "source_urls", 40),
        ("airtable review_status", "review_status", 16),
    ],
}

# ---------------------------------------------------------------------------
# Raw layers. One row per ITEM as it arrived, before clustering — the sheet a
# reviewer needs to answer "did the gate drop something it shouldn't have?"
# and "did these five articles really describe one action?".
#
# Raw rows carry `pillars` (the cheap Haiku gate's coarse guess) rather than
# `competency` / `relevance`, which are only assigned in the clustering pass.
# That difference is the point: comparing the two sheets shows where the second,
# stronger model overruled the first.
# ---------------------------------------------------------------------------
FEDERAL_RAW = {
    "table": "Federal Raw",
    "sheet": "Federal raw items",
    "date_field": "date",
    "relevance_field": "pillars",
    "columns": [
        ("date", "date", 11),
        ("lane", "lane", 16),
        ("source", "source", 24),
        ("branch", "branch", 10),
        ("agency", "agency", 18),
        ("instrument", "instrument_type", 20),
        ("instrument id", "instrument_id", 14),
        ("verification", "verification", 13),
        ("gate pillars", "pillars", 22),
        ("headline", "headline", 60),
        ("notes", "notes", 70),
        ("actor", "actor", 22),
        ("status", "status", 16),
        ("link", "source_urls", 40),
        ("primary document", "document_url", 40),
        ("ingested", "ingested_at", 20),
    ],
}

CONGRESS_RAW = {
    "table": "Congress Raw",
    "sheet": "Congress raw items",
    "date_field": "date",
    "relevance_field": "pillars",
    "columns": [
        ("date", "date", 11),
        ("committee", "committee", 16),
        ("chamber", "chamber", 9),
        ("party source", "party_source", 12),
        ("source", "source", 28),
        ("activity", "activity_type", 18),
        ("gate pillars", "pillars", 22),
        ("headline", "headline", 60),
        ("notes", "notes", 70),
        ("actor", "actor", 22),
        ("bills", "bill_refs", 16),
        ("status", "status", 14),
        ("link", "source_urls", 40),
        ("ingested", "ingested_at", 20),
    ],
}

READ_ME = {
    "federal": [
        ("Two layers",
         "The 'Federal events' sheet is the deduped layer — one row per action, with the "
         "agency's release, the Federal Register document, the GAO report and every "
         "trade-press write-up merged. 'Federal raw items' is what arrived before "
         "clustering, one row per item. Read the events sheet to judge our calls; read the "
         "raw sheet to check whether the gate dropped something, or whether items that "
         "look like one action really are one."),
        ("gate pillars vs competency",
         "The raw sheet's 'gate pillars' is a cheap first-pass guess from a small model. "
         "The events sheet's 'competency' and 'relevance' come from a second, stronger pass "
         "against the full rubric, which frequently overrules the first. Disagreement "
         "between the two sheets is expected, not a bug."),
        ("What this is",
         "One row per federal executive-branch action from the Recoding America Federal "
         "Capacity Tracker, machine-classified against our four competencies. The tracker "
         "runs daily; this is a snapshot."),
        ("What we need from you",
         "For each row: is it a real action (not just messaging), and is our competency and "
         "relevance call right? Use the verdict column; add a one-line correction where the "
         "call is wrong. Rows you don't touch are read as 'no opinion', not 'keep'."),
        ("How a row got here",
         "An item passes only if a concrete INSTRUMENT can be named — a numbered OMB memo, "
         "agency guidance, a rule, an executive order, a workforce or procurement action, a "
         "launch, a reorganisation, a report with findings, or a court order. Statements, "
         "ICYMI items, interviews and op-eds are dropped before this sheet."),
        ("lane",
         "executive-action = the agency's own instrument. news = trade press or The Hill "
         "reporting it. rulemaking = a Federal Register document. Lanes are ordered by "
         "provenance: where a cluster spans lanes, the instrument wins."),
        ("verification",
         "official = published by the agency, the Register, or the White House. reported = a "
         "credible outlet says it happened, no primary document. draft-leaked = reporting on "
         "a draft or an internal document. Reported items are kept on purpose — they are "
         "often the earliest signal — so judge them on the action, not the paperwork."),
        ("competency",
         "civil-service / procedure / digital / incentives, or blank. Blank is the common and "
         "correct outcome: most government activity is not about the government's own "
         "capacity. Direction is irrelevant — dismantling a capacity counts as much as "
         "building one."),
        ("relevance",
         "1-3, how central an example this is of the competencies listed. Blank when no "
         "competency matched."),
        ("Neutral restatement",
         "Titles and summaries are deliberately rewritten with promotional and partisan "
         "adjectives removed. If a summary reads flatter than the source, that is the "
         "intent — check it against the primary document column."),
    ],
    "congress": [
        ("Layers",
         "'Committee activity' is the deduped layer — one row per action, with majority and "
         "minority write-ups of the same thing merged. 'Congress raw items' is what arrived "
         "before clustering. Hearings and bills come from the Congress.gov API and need no "
         "clustering: one API record is one hearing or one bill."),
        ("gate pillars vs competency",
         "The raw sheet's 'gate pillars' is a cheap first-pass guess from a small model. "
         "The activity sheet's 'competency' and 'relevance' come from a second, stronger "
         "pass against the full rubric, which frequently overrules the first. Disagreement "
         "between the two sheets is expected, not a bug."),
        ("What this is",
         "Three sheets from the Recoding America Congressional Tracker: committee and member "
         "activity (clustered from press feeds), hearings, and bills. Hearings and bills come "
         "from the Congress.gov API; activity is scraped from committee and member press "
         "feeds. The tracker runs daily; this is a snapshot."),
        ("What we need from you",
         "For each row: is it a real action, and is our competency and relevance call right? "
         "Use the verdict column; add a one-line correction where the call is wrong. Rows you "
         "don't touch are read as 'no opinion', not 'keep'."),
        ("How a row got here",
         "For press items, an item must describe a concrete action — a letter sent, a markup "
         "held, a report released — not a reaction to one. Bills are everything the seven "
         "tracked committees acted on in the window, so most of the bill sheet is expected "
         "to be 'none'."),
        ("competency",
         "civil-service / procedure / digital / incentives, or blank, pointed at the FEDERAL "
         "government. Blank is the common and correct outcome. Two carve-outs differ from the "
         "state tracker: election administration counts (as digital or procedure), and "
         "appropriations are 'none' unless the funding MODEL changes."),
        ("relevance",
         "1-3, how central an example this is of the competencies listed. Blank when no "
         "competency matched."),
        ("GAO volume",
         "GAO reports are oversight-of-capacity almost by construction and outnumber the "
         "committees. They are in the activity sheet under committee 'gao'; filter them out "
         "if you want to review committee behaviour on its own."),
    ],
}


def write_readme(wb, kind, window_label, counts):
    ws = wb.create_sheet("Read me", 0)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 110
    ws["A1"] = f"{kind.title()} activity — review workbook"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Exported {date.today().isoformat()} · window: {window_label}"
    ws["A2"].font = Font(italic=True, color="475569")
    row = 4
    for sheet, n in counts.items():
        ws.cell(row=row, column=1, value=sheet).font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{n} rows")
        row += 1
    row += 1
    for label, text in READ_ME[kind]:
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="top")
        t = ws.cell(row=row, column=2, value=text)
        t.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(15, 13 * (len(text) // 95 + 1))
        row += 1
    return ws


def write_sheet(wb, spec, records):
    ws = wb.create_sheet(spec["sheet"])
    headers = [h for h, _, _ in spec["columns"]] + REVIEW_COLUMNS
    ws.append(headers)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26

    for r, rec in enumerate(records, start=2):
        f = rec["fields"]
        for i, (_, field, _) in enumerate(spec["columns"], start=1):
            raw = f.get(field)
            if field in ("source_urls", "document_url"):
                url = first_url(raw) if field == "source_urls" else (raw or "")
                # Rows written before the pipeline stopped pointing news items
                # at themselves carry document_url == the article link. An
                # identical "primary document" column is worse than a blank one.
                if field == "document_url" and url == first_url(f.get("source_urls")):
                    url = ""
                cell = ws.cell(row=r, column=i, value=url)
                if url:
                    cell.hyperlink = url
                    cell.font = Font(color="2563EB", underline="single", size=10)
                continue
            ws.cell(row=r, column=i, value=joined(raw))
        # Shade rows that matched no competency — they are legitimately here
        # (the reviewer is checking for false negatives) but they are not the
        # point of the sheet. Raw sheets test `pillars` instead: the clean
        # layer's `competency` field does not exist there.
        if not f.get(spec.get("relevance_field", "competency")):
            for i in range(1, len(headers) + 1):
                ws.cell(row=r, column=i).fill = NONE_FILL
        for i in range(len(spec["columns"]) + 1, len(headers) + 1):
            ws.cell(row=r, column=i).fill = REVIEW_FILL

    for i, (_, _, width) in enumerate(spec["columns"], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for i, _ in enumerate(REVIEW_COLUMNS, start=len(spec["columns"]) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    last_col = get_column_letter(len(headers))
    last_row = max(2, len(records) + 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = "A2"

    dv = DataValidation(type="list", formula1=f'"{",".join(VERDICT_CHOICES)}"', allow_blank=True)
    ws.add_data_validation(dv)
    verdict_col = get_column_letter(len(spec["columns"]) + 1)
    dv.add(f"{verdict_col}2:{verdict_col}{last_row}")
    return len(records)


def load(api, spec, cutoff, relevant_only):
    """Read a table, window it, and sort the way a reviewer wants to read it:
    most recent first, and within a day the strongest competency fit first."""
    try:
        table = api.base(AIRTABLE_BASE_ID).table(spec["table"])
        records = table.all()
    except Exception as e:
        print(f"  skipping {spec['table']}: {e}")
        return []
    out = []
    for rec in records:
        f = rec["fields"]
        if cutoff and (f.get(spec["date_field"]) or "") < cutoff:
            continue
        if relevant_only and not f.get(spec.get("relevance_field", "competency")):
            continue
        out.append(rec)
    out.sort(key=lambda r: ((r["fields"].get(spec["date_field"]) or ""),
                            r["fields"].get("relevance") or 0), reverse=True)
    return out


def build(api, kind, specs, cutoff, window_label, relevant_only, out_dir):
    wb = Workbook()
    wb.remove(wb.active)
    counts = {}
    for spec in specs:
        records = load(api, spec, cutoff, relevant_only)
        counts[spec["sheet"]] = write_sheet(wb, spec, records)
    write_readme(wb, kind, window_label, counts)
    wb.active = 0
    path = os.path.join(out_dir, f"{kind}-activity-{date.today().isoformat()}.xlsx")
    wb.save(path)
    total = sum(counts.values())
    print(f"  {path}  ({total} rows: "
          f"{', '.join(f'{n} {s.lower()}' for s, n in counts.items())})")
    return path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tracker", choices=["federal", "congress", "both"], default="both")
    ap.add_argument("--days", type=int, default=21)
    ap.add_argument("--all", action="store_true", help="ignore the window")
    ap.add_argument("--relevant-only", action="store_true",
                    help="drop rows that matched no competency")
    ap.add_argument("--out", default="review")
    args = ap.parse_args()

    if not (AIRTABLE_TOKEN and AIRTABLE_BASE_ID):
        sys.exit("Missing AIRTABLE_TOKEN / AIRTABLE_BASE_ID. See .env_example.")

    cutoff = "" if args.all else (date.today() - timedelta(days=args.days)).isoformat()
    window_label = "all time" if args.all else f"last {args.days} days (from {cutoff})"
    os.makedirs(args.out, exist_ok=True)
    api = Api(AIRTABLE_TOKEN)

    print(f"Exporting {args.tracker} · {window_label}"
          f"{' · RA-relevant only' if args.relevant_only else ''}")
    # Deduped layer first in each workbook — it is what the tab shows and what
    # most reviewers will read — then the raw layer behind it.
    if args.tracker in ("federal", "both"):
        build(api, "federal", [FEDERAL_EVENTS, FEDERAL_RAW], cutoff, window_label,
              args.relevant_only, args.out)
    if args.tracker in ("congress", "both"):
        build(api, "congress",
              [CONGRESS_EVENTS, CONGRESS_RAW, CONGRESS_HEARINGS, CONGRESS_BILLS],
              cutoff, window_label, args.relevant_only, args.out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
